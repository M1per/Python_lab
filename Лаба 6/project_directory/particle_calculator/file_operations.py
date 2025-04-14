# particle_calculator/file_operations.py

from docx import Document
import openpyxl

def save_to_docx(filename, data):
    """Save data to a Word document."""
    doc = Document()
    for key, value in data.items():
        doc.add_paragraph(f"{key}: {value}")
    doc.save(filename)

def save_to_xlsx(filename, data):
    """Save data to an Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for row, (key, value) in enumerate(data.items(), start=1):
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value)
    wb.save(filename)