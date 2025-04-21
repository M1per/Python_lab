from abc import ABC, abstractmethod
from docx import Document
import openpyxl


class FileSaver(ABC):
    @property
    @abstractmethod
    def default_ext(self):
        pass

    @property
    @abstractmethod
    def file_type(self):
        pass

    @abstractmethod
    def save(self, filename, data):
        pass


class DocxSaver(FileSaver):
    @property
    def default_ext(self):
        return ".docx"

    @property
    def file_type(self):
        return "Word Document"

    def save(self, filename, data):
        doc = Document()
        for key, value in data.items():
            doc.add_paragraph(f"{key}: {value}")
        doc.save(filename)


class XlsxSaver(FileSaver):
    @property
    def default_ext(self):
        return ".xlsx"

    @property
    def file_type(self):
        return "Excel Workbook"

    def save(self, filename, data):
        wb = openpyxl.Workbook()
        ws = wb.active
        for i, (key, value) in enumerate(data.items(), 1):
            ws.cell(row=i, column=1, value=key)
            ws.cell(row=i, column=2, value=value)
        wb.save(filename)